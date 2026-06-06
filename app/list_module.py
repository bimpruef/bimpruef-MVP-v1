"""
list_module.py – BIMPruef Element-Listen-Modul (Direct-from-Documents)

معماری جدید: بدون slot cache
- کاربر مستقیم از Documents فایل IFC انتخاب می‌کند
- فیلتر و ستون دلخواه تنظیم می‌کند
- دکمه «نمایش» → فایل از R2 لود، پردازش، JSON برمی‌گردد
- بدون session_id، بدون slot

Routen:
  GET  /projects/{project_id}/list              → Haupt-UI
  POST /projects/{project_id}/list/run          → JSON-API: Element laden + filtern
  GET  /projects/{project_id}/list/pset-keys    → Pset-Schlüssel für Datei
  GET  /projects/{project_id}/list/export       → Excel-Download
"""

import html as _html
import io
import json
import os
import tempfile
import zipfile

import ifcopenshell

from fastapi import APIRouter, Query, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response

from app.auth import require_user
from app.document_storage import (
    get_document,
    list_project_ifc_documents,
)
from app.extractors import (
    apply_filters as _apply_filters,
    extract_element_data,
    flatten_psets as _flatten_psets,
    get_candidate_products,
    get_psets_safe,
)
from app.project_storage import get_project

try:
    from app.r2_storage import download_file_from_r2, r2_enabled
except Exception:
    download_file_from_r2 = None
    r2_enabled = lambda: False

list_router = APIRouter()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _e(s) -> str:
    return _html.escape(str(s or ""))


def _fmt_size(num: int) -> str:
    try:
        n = float(num or 0)
    except Exception:
        return "0 B"
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024 or unit == "GB":
            return f"{n:.1f} {unit}" if unit != "B" else f"{int(n)} B"
        n /= 1024
    return f"{n:.1f} GB"


def _account_from_request(request: Request) -> dict:
    user = require_user(request)
    return {
        "account_id": user["user_id"],
        "account_name": user["email"],
        "workspace": "Personal",
    }


def _load_context(request: Request, project_id: str):
    account = _account_from_request(request)
    project = get_project(account["account_id"], project_id)
    if not project:
        return None, None
    return account, project


# ─────────────────────────────────────────────────────────────────────────────
# IFC از R2 لود کن (مثل project_clash.py)
# ─────────────────────────────────────────────────────────────────────────────

def _open_ifc_from_document(account_id: str, project_id: str, document_id: str):
    """
    IFC/IFCZIP را مستقیم از R2 لود می‌کند.
    Returns: (model, label)
    """
    if not (r2_enabled() and download_file_from_r2):
        raise ValueError("Cloudflare R2 ist nicht konfiguriert.")

    doc = get_document(account_id, project_id, document_id)
    label = doc.get("original_filename", document_id)
    ext = (doc.get("file_extension") or ".ifc").lower()

    fd, tmp_path = tempfile.mkstemp(prefix="bp_list_", suffix=ext)
    os.close(fd)

    try:
        download_file_from_r2(doc["r2_key"], tmp_path)

        if ext == ".ifczip":
            with open(tmp_path, "rb") as f:
                raw = f.read()
            with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
                ifc_names = [n for n in zf.namelist() if n.lower().endswith(".ifc")]
                if not ifc_names:
                    raise ValueError(f"Keine .ifc-Datei in '{label}' gefunden.")
                ifc_bytes = zf.read(ifc_names[0])

            fd2, ifc_tmp = tempfile.mkstemp(prefix="bp_list_ifc_", suffix=".ifc")
            os.close(fd2)
            with open(ifc_tmp, "wb") as f:
                f.write(ifc_bytes)
            try:
                model = ifcopenshell.open(ifc_tmp)
            finally:
                try:
                    os.remove(ifc_tmp)
                except OSError:
                    pass
        else:
            model = ifcopenshell.open(tmp_path)

    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return model, label


def _extract_and_filter(model, file_label: str, document_id: str,
                        filters: list, include_psets: bool) -> list:
    """المان‌ها را از مدل استخراج و فیلتر می‌کند."""
    all_elements = []
    for elem in get_candidate_products(model):
        if include_psets:
            data = extract_element_data(elem, file_label=file_label)
        else:
            data = {
                "file_label": file_label,
                "document_id": document_id,
                "express_id": elem.id() if hasattr(elem, "id") else "",
                "type": elem.is_a() if hasattr(elem, "is_a") else "",
                "name": getattr(elem, "Name", None) or "",
                "global_id": getattr(elem, "GlobalId", None) or "",
                "object_type": getattr(elem, "ObjectType", None) or "",
                "predefined_type": str(getattr(elem, "PredefinedType", None) or ""),
                "psets": {},
            }
        data["document_id"] = document_id
        all_elements.append(data)

    return _apply_filters(all_elements, filters)


def _collect_pset_keys_from_model(model) -> list:
    """همه Pset-کلیدهای یک مدل را جمع می‌کند."""
    keys: set[str] = set()
    for elem in get_candidate_products(model):
        psets = get_psets_safe(elem)
        for pset_name, props in (psets or {}).items():
            if isinstance(props, dict):
                for prop_name in props:
                    keys.add(f"pset:{pset_name}.{prop_name}")
    return sorted(keys)


# ─────────────────────────────────────────────────────────────────────────────
# JSON-API: المان‌ها لود و فیلتر کن
# ─────────────────────────────────────────────────────────────────────────────

@list_router.post("/projects/{project_id}/list/run")
async def list_run_api(request: Request, project_id: str):
    """
    Body (JSON):
    {
      "document_ids": ["id1", "id2", ...],
      "filters": [...],
      "columns": [...],
      "include_psets": true|false
    }
    """
    account, project = _load_context(request, project_id)
    if not project:
        return JSONResponse({"error": "Projekt nicht gefunden."}, status_code=404)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "Ungültiger JSON-Body."}, status_code=400)

    document_ids = [str(d).strip() for d in body.get("document_ids", []) if str(d).strip()]
    filters = body.get("filters", [])
    columns = body.get("columns", [])
    include_psets = bool(body.get("include_psets", False))

    # بررسی آیا pset در فیلتر یا ستون هست
    needs_psets = include_psets or any(
        (isinstance(f, dict) and str(f.get("field", "")).startswith("pset:"))
        for f in filters
    ) or any(
        isinstance(c, str) and c.startswith("pset:")
        for c in columns
    )

    if not document_ids:
        return JSONResponse({"error": "Bitte mindestens eine Datei auswählen."}, status_code=400)

    all_rows = []
    all_pset_keys: set[str] = set()
    file_errors = []

    for doc_id in document_ids:
        try:
            model, label = _open_ifc_from_document(
                account["account_id"], project_id, doc_id
            )
        except Exception as exc:
            file_errors.append(f"{doc_id}: {exc}")
            continue

        filtered = _extract_and_filter(model, label, doc_id, filters, needs_psets)

        if needs_psets:
            for elem in filtered:
                for k in _flatten_psets(elem.get("psets", {})).keys():
                    all_pset_keys.add(f"pset:{k}")

        # ردیف‌های خروجی
        selected_pset_cols = [c for c in columns if isinstance(c, str) and c.startswith("pset:")]

        for elem in filtered:
            row = {
                "file_label": elem.get("file_label", ""),
                "document_id": elem.get("document_id", ""),
                "express_id": elem.get("express_id", ""),
                "type": elem.get("type", ""),
                "name": elem.get("name", ""),
                "global_id": elem.get("global_id", ""),
                "object_type": elem.get("object_type", ""),
                "predefined_type": elem.get("predefined_type", ""),
            }
            if needs_psets and selected_pset_cols:
                flat = _flatten_psets(elem.get("psets", {}))
                for col_key in selected_pset_cols:
                    row[col_key] = flat.get(col_key[5:], "")
            all_rows.append(row)

    return JSONResponse({
        "total": len(all_rows),
        "pset_keys": sorted(all_pset_keys),
        "rows": all_rows,
        "file_errors": file_errors,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Pset-کلیدها برای یک فایل (برای Filter-Dropdown)
# ─────────────────────────────────────────────────────────────────────────────

@list_router.get("/projects/{project_id}/list/pset-keys")
def list_pset_keys(
    request: Request,
    project_id: str,
    document_id: str = Query(...),
):
    account, project = _load_context(request, project_id)
    if not project:
        return JSONResponse({"error": "Projekt nicht gefunden."}, status_code=404)

    try:
        model, _ = _open_ifc_from_document(account["account_id"], project_id, document_id)
        keys = _collect_pset_keys_from_model(model)
        return JSONResponse({"pset_keys": keys})
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
# Excel Export
# ─────────────────────────────────────────────────────────────────────────────

@list_router.post("/projects/{project_id}/list/export")
async def list_export_excel(request: Request, project_id: str):
    """همان منطق /run اما خروجی Excel."""
    account, project = _load_context(request, project_id)
    if not project:
        return Response(content="Projekt nicht gefunden.", status_code=404)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(content="openpyxl nicht installiert.", status_code=500)

    try:
        body = await request.json()
    except Exception:
        return Response(content="Ungültiger JSON-Body.", status_code=400)

    document_ids = [str(d).strip() for d in body.get("document_ids", []) if str(d).strip()]
    filters = body.get("filters", [])
    columns = body.get("columns", [])

    needs_psets = any(
        (isinstance(f, dict) and str(f.get("field", "")).startswith("pset:"))
        for f in filters
    ) or any(isinstance(c, str) and c.startswith("pset:") for c in columns)

    BASE_COLUMNS = [
        ("file_label", "Datei"),
        ("express_id", "Express-ID"),
        ("type", "IFC-Typ"),
        ("name", "Name"),
        ("global_id", "GlobalId"),
        ("object_type", "ObjectType"),
        ("predefined_type", "PredefinedType"),
    ]
    base_col_keys = {k for k, _ in BASE_COLUMNS}

    if columns:
        export_cols = []
        for col_key in columns:
            if col_key in base_col_keys:
                label = next((lbl for k, lbl in BASE_COLUMNS if k == col_key), col_key)
                export_cols.append((col_key, label))
            elif col_key.startswith("pset:"):
                export_cols.append((col_key, col_key[5:]))
    else:
        export_cols = BASE_COLUMNS

    all_rows = []
    for doc_id in document_ids:
        try:
            model, label = _open_ifc_from_document(account["account_id"], project_id, doc_id)
            filtered = _extract_and_filter(model, label, doc_id, filters, needs_psets)
            selected_pset_cols = [c for c in columns if isinstance(c, str) and c.startswith("pset:")]
            for elem in filtered:
                row = {k: elem.get(k, "") for k, _ in BASE_COLUMNS}
                if needs_psets and selected_pset_cols:
                    flat = _flatten_psets(elem.get("psets", {}))
                    for col_key in selected_pset_cols:
                        row[col_key] = flat.get(col_key[5:], "")
                all_rows.append(row)
        except Exception:
            continue

    # Excel باسازی
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BIMPruef Elementliste"

    header_fill  = PatternFill("solid", fgColor="0F2040")
    header_font  = Font(name="Calibri", bold=True, color="4FC3F7", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side    = Side(style="thin", color="1E3A6E")
    cell_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    alt_fill     = PatternFill("solid", fgColor="1A2A4A")
    normal_fill  = PatternFill("solid", fgColor="16213E")
    cell_font    = Font(name="Calibri", color="D0DCE8", size=10)
    cell_align   = Alignment(vertical="top", wrap_text=False)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(export_cols), 1))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"BIMPruef – Elementliste  |  {len(all_rows)} Elemente  |  Projekt: {project.get('project_name', '')}"
    title_cell.font = Font(name="Calibri", bold=True, color="4FC3F7", size=13)
    title_cell.fill = PatternFill("solid", fgColor="0E0E1A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    for col_idx, (col_key, col_label) in enumerate(export_cols, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = col_label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
    ws.row_dimensions[2].height = 22

    for row_idx, row_data in enumerate(all_rows, start=3):
        fill = alt_fill if row_idx % 2 == 0 else normal_fill
        for col_idx, (col_key, _) in enumerate(export_cols, start=1):
            value = row_data.get(col_key, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = str(value) if value is not None else ""
            cell.font = cell_font
            cell.fill = fill
            cell.alignment = cell_align
            cell.border = cell_border

    for col_idx, (_, col_label) in enumerate(export_cols, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(col_label)
        for row_idx in range(3, min(3 + len(all_rows), 203)):
            v = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A3"
    if export_cols:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(export_cols))}2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="bimpruef_elementliste.xlsx"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Haupt-UI
# ─────────────────────────────────────────────────────────────────────────────

@list_router.get("/projects/{project_id}/list", response_class=HTMLResponse)
def project_list(request: Request, project_id: str):
    try:
        account, project = _load_context(request, project_id)
    except Exception:
        return RedirectResponse("/auth/login", status_code=302)

    if not project:
        return RedirectResponse("/", status_code=302)

    account_id = account["account_id"]
    docs = list_project_ifc_documents(account_id, project_id)

    return _render_list_page(account, project, docs)


# ─────────────────────────────────────────────────────────────────────────────
# Legacy redirects
# ─────────────────────────────────────────────────────────────────────────────

@list_router.get("/viewer/list/", response_class=HTMLResponse)
def viewer_list_legacy(project_id: str = Query(default="")):
    if project_id:
        return RedirectResponse(f"/projects/{project_id}/list", status_code=302)
    return RedirectResponse("/", status_code=302)


@list_router.get("/viewer/list/data/", response_class=HTMLResponse)
def viewer_list_data_legacy():
    return JSONResponse(
        {"error": "Dieser Endpunkt wurde nach /projects/{project_id}/list/run verschoben."},
        status_code=410,
    )


@list_router.get("/viewer/list/export/", response_class=HTMLResponse)
def viewer_list_export_legacy():
    return Response(
        content="Dieser Endpunkt wurde nach /projects/{project_id}/list/export verschoben.",
        status_code=410,
    )


# ─────────────────────────────────────────────────────────────────────────────
# HTML رندر
# ─────────────────────────────────────────────────────────────────────────────

def _render_list_page(account: dict, project: dict, docs: list) -> HTMLResponse:
    from app.projects import _page, _project_subnav, _topbar_global

    project_id = project["project_id"]
    pid = _e(project_id)

    # اگر هیچ فایلی نیست
    if not docs:
        body = f"""
        {_topbar_global(account)}
        {_project_subnav(project_id, "list")}
        <div style="padding:28px 32px;max-width:900px;margin:0 auto">
          <h1 style="font-size:22px;font-weight:600;margin-bottom:14px">Elementliste</h1>
          <div class="card" style="border-color:var(--accent2)">
            <h3 style="font-size:15px;margin-bottom:8px">Keine IFC-Dateien im Documents-Modul</h3>
            <p style="color:var(--muted);font-size:13px;margin-bottom:14px">
              Das List-Modul lädt IFC-Dateien direkt aus dem Documents-Modul.
              Lade zuerst mindestens eine .ifc- oder .ifczip-Datei dort hoch.
            </p>
            <a class="btn btn-primary" href="/projects/{pid}/documents"
               style="text-decoration:none">Zu Documents</a>
          </div>
        </div>"""
        return _page(f"{project['project_name']} – Liste", body)

    # گزینه‌های فایل برای انتخاب
    file_checkboxes = ""
    for i, d in enumerate(docs):
        size_label = _fmt_size(d.get("file_size", 0))
        folder = _e(d.get("folder_path") or "Root")
        doc_id = _e(d["document_id"])
        fname = _e(d["original_filename"])
        checked = "checked" if i == 0 else ""
        file_checkboxes += f"""
        <label style="display:flex;align-items:center;gap:8px;padding:8px 10px;
          border-radius:7px;cursor:pointer;font-size:12px;
          background:var(--surface2);border:1px solid var(--border);
          transition:border-color .15s"
          onmouseenter="this.style.borderColor='var(--accent)'"
          onmouseleave="this.style.borderColor='var(--border)'">
          <input type="checkbox" class="doc-chk" value="{doc_id}" {checked}
            style="accent-color:var(--accent);width:13px;height:13px;flex-shrink:0">
          <div style="flex:1;min-width:0">
            <div style="font-weight:600;color:var(--text);overflow:hidden;
              text-overflow:ellipsis;white-space:nowrap" title="{fname}">{fname}</div>
            <div style="font-size:10px;color:var(--muted);margin-top:1px">
              {size_label} &nbsp;·&nbsp; {folder}
            </div>
          </div>
        </label>"""

    body = f"""
{_topbar_global(account)}
{_project_subnav(project_id, "list")}

<div style="display:flex;flex-direction:column;height:calc(100vh - 94px);overflow:hidden">
  <div style="display:flex;flex:1;overflow:hidden">

    <!-- ── لفت پنل: کنترل‌ها ──────────────────────────────────────────── -->
    <div id="left-panel" style="width:360px;min-width:320px;background:var(--surface);
      border-right:1px solid var(--border);display:flex;flex-direction:column;
      overflow:hidden;flex-shrink:0">

      <div style="padding:8px 12px;font-size:10px;font-weight:700;
        background:var(--surface2);color:var(--muted);
        text-transform:uppercase;letter-spacing:.7px;
        display:flex;align-items:center;justify-content:space-between;flex-shrink:0;
        border-bottom:1px solid var(--border)">
        <span>📋 Elementliste</span>
        <button id="btn-run" class="btn btn-primary"
          style="font-size:11px;padding:3px 12px">▶ Laden</button>
      </div>

      <div style="flex:1;overflow-y:auto;padding:10px">

        <!-- فایل‌های IFC -->
        <div class="card" style="margin-bottom:10px">
          <div style="font-size:11px;font-weight:600;color:var(--muted);margin-bottom:8px;
            text-transform:uppercase;letter-spacing:.5px;
            display:flex;align-items:center;justify-content:space-between">
            <span>📁 IFC-Dateien</span>
            <a href="/projects/{pid}/documents"
              style="font-size:10px;color:var(--accent);text-decoration:none">Documents →</a>
          </div>
          <div style="display:flex;flex-direction:column;gap:5px">
            {file_checkboxes}
          </div>
        </div>

        <!-- فیلترها -->
        <div class="card" style="margin-bottom:10px">
          <div style="display:flex;align-items:center;justify-content:space-between;
            margin-bottom:8px">
            <div style="font-size:11px;font-weight:600;color:var(--muted);
              text-transform:uppercase;letter-spacing:.5px">⚙ Filter-Regeln</div>
            <button id="btn-add-filter" class="btn"
              style="font-size:10px;padding:2px 9px;color:var(--accent)">
              + Hinzufügen
            </button>
          </div>
          <div id="filter-list" style="display:flex;flex-direction:column;gap:6px"></div>
          <div id="no-filters-hint" style="font-size:11px;color:var(--muted);
            font-style:italic;padding:4px 0">
            Kein Filter aktiv – alle Elemente werden angezeigt.
          </div>
        </div>

        <!-- ستون‌ها -->
        <div class="card" style="margin-bottom:10px">
          <div style="display:flex;align-items:center;justify-content:space-between;
            margin-bottom:8px">
            <div style="font-size:11px;font-weight:600;color:var(--muted);
              text-transform:uppercase;letter-spacing:.5px">📊 Spalten</div>
            <div style="display:flex;gap:5px;flex-wrap:wrap;justify-content:flex-end">
              <button id="btn-load-psets" class="btn"
                style="font-size:10px;padding:2px 8px;color:var(--accent)">
                Pset-Spalten laden
              </button>
              <button id="btn-cols-all" class="btn"
                style="font-size:10px;padding:2px 8px">Alle</button>
              <button id="btn-cols-none" class="btn"
                style="font-size:10px;padding:2px 8px">Keine</button>
            </div>
          </div>
          <div id="column-list" style="display:flex;flex-direction:column;gap:4px">
            <div style="font-size:11px;color:var(--muted);font-style:italic">
              Lade Spaltenliste …
            </div>
          </div>
        </div>

      </div>

      <!-- دکمه Export -->
      <div style="padding:10px 12px;border-top:1px solid var(--border);flex-shrink:0">
        <button id="btn-export" class="btn btn-primary"
          style="width:100%;font-size:13px;padding:9px">
          ⬇ Als Excel herunterladen
        </button>
      </div>

    </div>

    <!-- ── راست: جدول نتایج ───────────────────────────────────────────── -->
    <div style="flex:1;display:flex;flex-direction:column;overflow:hidden">

      <div id="status-bar" style="padding:6px 14px;background:var(--surface2);font-size:11px;
        color:var(--muted);border-bottom:1px solid var(--border);flex-shrink:0;
        display:flex;align-items:center;gap:12px">
        <span id="status-total">Dateien auswählen und ▶ Laden klicken</span>
        <span id="status-filtered" style="color:var(--accent)"></span>
        <span style="margin-left:auto;color:var(--muted);font-size:10px" id="status-cols"></span>
      </div>

      <div id="table-wrap" style="flex:1;overflow:auto">
        <div style="display:flex;flex-direction:column;align-items:center;
          justify-content:center;height:100%;gap:14px;padding:40px"
          id="table-placeholder">
          <svg width="40" height="40" fill="none" stroke="var(--muted)" stroke-width="1.5" viewBox="0 0 24 24">
            <rect x="3" y="3" width="18" height="18" rx="2"/><path d="M3 9h18M9 21V9"/>
          </svg>
          <div style="text-align:center">
            <div style="font-size:14px;color:var(--text);margin-bottom:6px">
              Elementliste bereit
            </div>
            <div style="font-size:12px;color:var(--muted)">
              IFC-Datei(en) wählen, optional Filter setzen,<br>
              dann <strong style="color:var(--accent)">▶ Laden</strong> klicken.
            </div>
          </div>
        </div>
        <table id="result-table" style="display:none">
          <thead id="result-thead"></thead>
          <tbody id="result-tbody"></tbody>
        </table>
      </div>

    </div>
  </div>
</div>

<script>
(function() {{

const PROJECT_ID  = {json.dumps(project_id)};
const RUN_URL     = `/projects/${{encodeURIComponent(PROJECT_ID)}}/list/run`;
const PSET_URL    = `/projects/${{encodeURIComponent(PROJECT_ID)}}/list/pset-keys`;
const EXPORT_URL  = `/projects/${{encodeURIComponent(PROJECT_ID)}}/list/export`;

let allRows       = [];
let psetKeys      = [];
let filters       = [];
let selectedCols  = [];
let colDefs       = [];
let filterCounter = 0;

const BASE_COLS = [
  {{key:"file_label",      label:"Datei"}},
  {{key:"express_id",      label:"Express-ID"}},
  {{key:"type",            label:"IFC-Typ"}},
  {{key:"name",            label:"Name"}},
  {{key:"global_id",       label:"GlobalId"}},
  {{key:"object_type",     label:"ObjectType"}},
  {{key:"predefined_type", label:"PredefinedType"}},
];

const BASE_FIELD_OPTS = [
  {{key:"file_label",      label:"Datei"}},
  {{key:"type",            label:"IFC-Typ"}},
  {{key:"name",            label:"Name"}},
  {{key:"global_id",       label:"GlobalId"}},
  {{key:"object_type",     label:"ObjectType"}},
  {{key:"predefined_type", label:"PredefinedType"}},
];
const OPERATORS = [
  {{key:"contains",     label:"enthält"}},
  {{key:"not_contains", label:"enthält nicht"}},
  {{key:"equals",       label:"ist gleich"}},
  {{key:"not_equals",   label:"ist ungleich"}},
  {{key:"starts_with",  label:"beginnt mit"}},
  {{key:"ends_with",    label:"endet mit"}},
];

const filterList       = document.getElementById("filter-list");
const noFiltersHint    = document.getElementById("no-filters-hint");
const columnList       = document.getElementById("column-list");
const resultTable      = document.getElementById("result-table");
const resultThead      = document.getElementById("result-thead");
const resultTbody      = document.getElementById("result-tbody");
const tablePlaceholder = document.getElementById("table-placeholder");
const statusTotal      = document.getElementById("status-total");
const statusFiltered   = document.getElementById("status-filtered");
const statusCols       = document.getElementById("status-cols");
const btnRun           = document.getElementById("btn-run");
const btnExport        = document.getElementById("btn-export");
const btnLoadPsets     = document.getElementById("btn-load-psets");

function esc(s) {{
  return String(s||"").replace(/&/g,"&amp;").replace(/</g,"&lt;")
    .replace(/>/g,"&gt;").replace(/"/g,"&quot;");
}}

// ── فیلدهای فیلتر ────────────────────────────────────────────────────────
function buildFieldOptions(extraPsetKeys) {{
  let opts = BASE_FIELD_OPTS.map(f =>
    `<option value="${{f.key}}">${{esc(f.label)}}</option>`
  ).join("");
  if (extraPsetKeys.length) {{
    opts += `<optgroup label="Eigenschaften (Psets)">`;
    for (const k of extraPsetKeys) {{
      const lbl = k.startsWith("pset:") ? k.slice(5) : k;
      opts += `<option value="${{esc(k)}}">${{esc(lbl)}}</option>`;
    }}
    opts += `</optgroup>`;
  }}
  return opts;
}}

// ── فیلتر اضافه کردن ─────────────────────────────────────────────────────
function addFilter(fieldKey, operator, value) {{
  const id = ++filterCounter;
  const wrapper = document.createElement("div");
  wrapper.id = `filter-${{id}}`;
  wrapper.style.cssText = "display:grid;gap:4px;padding:8px;background:var(--bg);" +
    "border:1px solid var(--border);border-radius:6px";

  const fieldOpts = buildFieldOptions(psetKeys);
  const opOpts = OPERATORS.map(o =>
    `<option value="${{o.key}}" ${{o.key===operator?"selected":""}}>${{o.label}}</option>`
  ).join("");

  wrapper.innerHTML = `
    <div style="display:flex;align-items:center;gap:4px">
      <select class="filter-field" data-id="${{id}}"
        style="flex:1;background:var(--bg);border:1px solid var(--border);color:var(--text);
        padding:4px 6px;border-radius:5px;font-size:11px">
        ${{fieldOpts}}
      </select>
      <button class="btn-remove-filter" data-id="${{id}}"
        style="background:var(--surface2);border:1px solid var(--accent2);color:var(--accent2);
        padding:2px 8px;border-radius:4px;font-size:11px;cursor:pointer;flex-shrink:0">✕</button>
    </div>
    <select class="filter-op" data-id="${{id}}"
      style="background:var(--bg);border:1px solid var(--border);color:var(--text);
      padding:4px 6px;border-radius:5px;font-size:11px">
      ${{opOpts}}
    </select>
    <input class="filter-val" data-id="${{id}}" type="text"
      value="${{esc(value||"")}}"
      placeholder="Suchwert eingeben …"
      style="background:var(--bg);border:1px solid var(--border);color:var(--text);
      padding:4px 8px;border-radius:5px;font-size:11px;width:100%;outline:none">
  `;

  if (fieldKey) wrapper.querySelector(".filter-field").value = fieldKey;

  filterList.appendChild(wrapper);
  noFiltersHint.style.display = "none";

  wrapper.querySelector(".btn-remove-filter").addEventListener("click", () => {{
    wrapper.remove();
    if (!filterList.children.length) noFiltersHint.style.display = "";
    syncFilters();
  }});

  syncFilters();
}}

function syncFilters() {{
  filters = [];
  document.querySelectorAll("#filter-list > div").forEach(row => {{
    const f = row.querySelector(".filter-field");
    const o = row.querySelector(".filter-op");
    const v = row.querySelector(".filter-val");
    if (f && o && v) {{
      filters.push({{field: f.value, operator: o.value, value: v.value}});
    }}
  }});
}}

// ── ستون‌ها ───────────────────────────────────────────────────────────────
function buildColumnUI() {{
  colDefs = [...BASE_COLS];
  psetKeys.forEach(k => {{
    const lbl = k.startsWith("pset:") ? k.slice(5) : k;
    colDefs.push({{key: k, label: lbl}});
  }});

  columnList.innerHTML = "";

  const baseSection = document.createElement("div");
  baseSection.style.cssText = "margin-bottom:8px";
  baseSection.innerHTML = `<div style="font-size:10px;color:var(--muted);margin-bottom:4px;
    text-transform:uppercase;letter-spacing:.4px">Basis-Felder</div>`;
  BASE_COLS.forEach(col => baseSection.appendChild(makeColRow(col)));
  columnList.appendChild(baseSection);

  if (psetKeys.length) {{
    const psetSection = document.createElement("div");
    psetSection.innerHTML = `<div style="font-size:10px;color:var(--muted);margin-bottom:4px;
      text-transform:uppercase;letter-spacing:.4px">Eigenschaften (Psets)</div>`;
    psetKeys.forEach(k => {{
      const lbl = k.startsWith("pset:") ? k.slice(5) : k;
      psetSection.appendChild(makeColRow({{key:k, label:lbl}}));
    }});
    columnList.appendChild(psetSection);
  }}

  syncColumns();
}}

function makeColRow(col) {{
  const div = document.createElement("label");
  div.style.cssText = "display:flex;align-items:center;gap:6px;cursor:pointer;" +
    "padding:3px 4px;border-radius:4px;font-size:11px";
  div.innerHTML = `
    <input type="checkbox" class="col-chk" value="${{esc(col.key)}}" checked
      style="accent-color:var(--accent);width:12px;height:12px;flex-shrink:0">
    <span style="color:var(--text);flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"
      title="${{esc(col.label)}}">${{esc(col.label)}}</span>
  `;
  div.addEventListener("mouseenter", () => div.style.background = "rgba(79,195,247,.06)");
  div.addEventListener("mouseleave", () => div.style.background = "");
  div.querySelector(".col-chk").addEventListener("change", syncColumns);
  return div;
}}

function syncColumns() {{
  selectedCols = [...document.querySelectorAll(".col-chk:checked")].map(c => c.value);
  statusCols.textContent = selectedCols.length + " Spalten ausgewählt";
}}

// ── لود داده (POST به /run) ───────────────────────────────────────────────
async function loadData() {{
  syncFilters();
  syncColumns();

  const docIds = [...document.querySelectorAll(".doc-chk:checked")].map(c => c.value);
  if (!docIds.length) {{
    statusTotal.textContent = "⚠ Bitte mindestens eine Datei auswählen.";
    return;
  }}

  btnRun.disabled = true;
  btnRun.textContent = "⏳ …";
  statusTotal.textContent = "Lade Daten aus R2 …";
  statusFiltered.textContent = "";

  const needsPsets = selectedCols.some(c => c.startsWith("pset:")) ||
                     filters.some(f => f.field && f.field.startsWith("pset:"));

  try {{
    const resp = await fetch(RUN_URL, {{
      method: "POST",
      headers: {{"Content-Type": "application/json"}},
      body: JSON.stringify({{
        document_ids: docIds,
        filters,
        columns: selectedCols,
        include_psets: needsPsets,
      }}),
    }});
    const data = await resp.json();
    if (data.error) throw new Error(data.error);

    allRows = data.rows || [];

    // Pset-Schlüssel aktualisieren falls zurückgegeben
    if (data.pset_keys && data.pset_keys.length && !psetKeys.length) {{
      psetKeys = data.pset_keys;
      buildColumnUI();
      // Filter-Dropdowns aktualisieren
      document.querySelectorAll(".filter-field").forEach(sel => {{
        const cur = sel.value;
        sel.innerHTML = buildFieldOptions(psetKeys);
        sel.value = cur;
      }});
    }}

    statusTotal.textContent = `${{allRows.length}} Elemente geladen`;
    if (data.file_errors && data.file_errors.length) {{
      statusFiltered.textContent = `⚠ ${{data.file_errors.length}} Datei-Fehler`;
    }} else {{
      statusFiltered.textContent = "";
    }}

    syncColumns();
    renderTable(allRows);

  }} catch(e) {{
    statusTotal.textContent = "Fehler: " + e.message;
    tablePlaceholder.style.display = "flex";
    tablePlaceholder.innerHTML =
      `<span style="color:var(--accent2)">⚠ ${{esc(e.message)}}</span>`;
    resultTable.style.display = "none";
  }} finally {{
    btnRun.disabled = false;
    btnRun.textContent = "▶ Laden";
  }}
}}

// ── جدول رندر ─────────────────────────────────────────────────────────────
function renderTable(rows) {{
  const activeCols = colDefs.filter(c => selectedCols.includes(c.key));
  if (!activeCols.length) {{
    tablePlaceholder.style.display = "flex";
    tablePlaceholder.innerHTML = "<span>Bitte mindestens eine Spalte auswählen.</span>";
    resultTable.style.display = "none";
    return;
  }}

  const visibleCols = activeCols.filter(c => {{
    return rows.some(row => {{
      const val = row[c.key];
      return val !== undefined && val !== null && String(val).trim() !== "";
    }});
  }});

  if (!visibleCols.length && rows.length) {{
    // همه ستون‌ها را نشان بده
    const allActiveCols = activeCols;
    _renderTableWith(rows, allActiveCols);
    return;
  }}

  _renderTableWith(rows, visibleCols.length ? visibleCols : activeCols);
}}

function _renderTableWith(rows, cols) {{
  let thead = "<tr>";
  thead += `<th style="min-width:40px;width:40px">#</th>`;
  cols.forEach(c => {{
    thead += `<th style="min-width:80px">${{esc(c.label)}}</th>`;
  }});
  thead += "</tr>";
  resultThead.innerHTML = thead;

  const MAX_PREVIEW = 2000;
  let tbody = "";
  rows.slice(0, MAX_PREVIEW).forEach((row, idx) => {{
    tbody += `<tr>`;
    tbody += `<td style="color:var(--muted);text-align:right">${{idx+1}}</td>`;
    cols.forEach(c => {{
      const val = row[c.key] !== undefined ? row[c.key] : "";
      tbody += `<td title="${{esc(String(val))}}">${{esc(String(val))}}</td>`;
    }});
    tbody += `</tr>`;
  }});
  if (rows.length > MAX_PREVIEW) {{
    tbody += `<tr><td colspan="${{cols.length+1}}"
      style="text-align:center;color:var(--muted);font-style:italic;padding:10px">
      … ${{rows.length - MAX_PREVIEW}} weitere Zeilen (im Excel-Export enthalten)
      </td></tr>`;
  }}

  resultTbody.innerHTML = tbody;
  tablePlaceholder.style.display = "none";
  resultTable.style.display = "";
}}

// ── Pset-کلیدها لود کن ───────────────────────────────────────────────────
async function loadPsetKeys() {{
  const docIds = [...document.querySelectorAll(".doc-chk:checked")].map(c => c.value);
  if (!docIds.length) {{
    statusTotal.textContent = "⚠ Bitte zuerst eine Datei auswählen.";
    return;
  }}

  btnLoadPsets.disabled = true;
  btnLoadPsets.textContent = "⏳ Psets …";

  const allKeys = new Set();
  let anyError = false;

  for (const docId of docIds) {{
    try {{
      const resp = await fetch(PSET_URL + `?document_id=${{encodeURIComponent(docId)}}`);
      const data = await resp.json();
      if (data.error) throw new Error(data.error);
      (data.pset_keys || []).forEach(k => allKeys.add(k));
    }} catch (e) {{
      anyError = true;
    }}
  }}

  psetKeys = [...allKeys].sort();
  buildColumnUI();

  document.querySelectorAll(".filter-field").forEach(sel => {{
    const cur = sel.value;
    sel.innerHTML = buildFieldOptions(psetKeys);
    sel.value = cur;
  }});

  btnLoadPsets.textContent = "Pset-Spalten laden";
  btnLoadPsets.disabled = false;
  statusTotal.textContent = `${{psetKeys.length}} Pset-Schlüssel geladen${{anyError ? " (mit Fehlern)" : ""}}`;
}}

// ── Excel Export ──────────────────────────────────────────────────────────
async function doExport() {{
  syncFilters();
  syncColumns();

  const docIds = [...document.querySelectorAll(".doc-chk:checked")].map(c => c.value);
  if (!docIds.length) {{
    alert("Bitte mindestens eine Datei auswählen.");
    return;
  }}

  btnExport.disabled = true;
  btnExport.textContent = "⏳ Exportiere …";

  try {{
    const resp = await fetch(EXPORT_URL, {{
      method: "POST",
      headers: {{"Content-Type": "application/json"}},
      body: JSON.stringify({{
        document_ids: docIds,
        filters,
        columns: selectedCols,
      }}),
    }});

    if (!resp.ok) throw new Error(`HTTP ${{resp.status}}`);

    const blob = await resp.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url;
    a.download = "bimpruef_elementliste.xlsx";
    document.body.appendChild(a);
    a.click();
    document.body.removeChild(a);
    URL.revokeObjectURL(url);

  }} catch(e) {{
    alert("Export-Fehler: " + e.message);
  }} finally {{
    btnExport.disabled = false;
    btnExport.textContent = "⬇ Als Excel herunterladen";
  }}
}}

// ── Event Listeners ───────────────────────────────────────────────────────
document.getElementById("btn-cols-all").addEventListener("click", () => {{
  document.querySelectorAll(".col-chk").forEach(c => {{ c.checked = true; }});
  syncColumns();
}});
document.getElementById("btn-cols-none").addEventListener("click", () => {{
  document.querySelectorAll(".col-chk").forEach(c => {{ c.checked = false; }});
  syncColumns();
}});

document.getElementById("btn-add-filter").addEventListener("click",
  () => addFilter("type", "contains", ""));
btnRun.addEventListener("click", loadData);
btnExport.addEventListener("click", doExport);
btnLoadPsets.addEventListener("click", loadPsetKeys);

document.addEventListener("keydown", e => {{
  if (e.key === "Enter" && e.target.classList.contains("filter-val")) loadData();
}});

// تغییر انتخاب فایل → Pset‌ها را ریست کن
document.querySelectorAll(".doc-chk").forEach(cb => {{
  cb.addEventListener("change", () => {{
    // اگر Pset‌های قدیمی مال فایل دیگری بود ریست نکن - فقط hint
  }});
}});

// ── اول بار ──────────────────────────────────────────────────────────────
buildColumnUI();

}})();
</script>
"""
    return _page(f"{project['project_name']} – Liste", body)
